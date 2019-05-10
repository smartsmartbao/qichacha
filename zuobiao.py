import requests
from openpyxl import *
import json
import csv
import math
import time
GAODE_API = 'http://restapi.amap.com/v3/geocode/geo?address={}&output=json&key=2df43ae9994dfea57373568e7aca7881'
column = ['address', 'lng', 'lat']
path = 'address1.xlsx'
FLAG = 1
#顺序省市区、citycode、adcode，gaode_lng,gaode_lat,wgs_lng,wsg_lat
class district(object):
    def __init__(self,path):
        self.path=path
        self.wb = load_workbook(self.path)
        self.wls =self.wb['Sheet0']
        self.row=self.wls.max_row #行数
        self.column=self.wls.max_column #列数
    def read_excel(self, dir):#读取excel的地址字段
        column = input("请输入在那一列例如A：")
        address =self.wls[column+'1':column+'10'] #有多少行写多少个
        return address
    def write_excel(self, dir, row, co, body): #row代表行,column代表列
        wb=load_workbook(dir)
        wls=wb['Sheet0']
        wls.cell(row = row, column= co, value=body)
        t=wls.cell(row = row, column= co)
        print (t.value)
        wb.save(dir)
    def sheng_shi_qu(self,address,row):
        b=[]
        try:
            r=json.loads(requests.get(GAODE_API.format(address)).text)
            location=r["geocodes"][0]
            province = location['province']
            city = location['city']
            district=location['district']
            b.append(province)
            b.append(city)
            b.append(district)
            co=self.column+1
            for i in range(len(b)):
                self.write_excel(self.path,row,co,b[i])
                co+=1

        except Exception as e:
            print(e)
    def runall(self,row):
        key=self.read_excel(self.path)
        for i in key:
            address=i[0].value
            self.sheng_shi_qu(address,row)
            row+=1
class gaode_code(object):
    def __init__(self,path):
        self.path=path
        self.wb=load_workbook(self.path)
        self.wls=self.wb['Sheet0']
        self.row=self.wls.max_row
        self.column=self.wls.max_column
    def read_excel(self):#读取excel的地址字段
        column = input("请输入在那一列例如A：")
        address =self.wls[column+'1':column+'10'] #有多少行写多少个
        return address
    def write_excel(self, dir, row, co, body):
        try: #row代表行,column代表列
            wb=load_workbook(dir)
            wls=wb['Sheet0']
            wls.cell(row = row, column= co, value=body)
            t=wls.cell(row = row, column= co)
            wb.save(dir)
            print (t.value)
        except Exception as e:
            print(e)
    def city_ad(self,address,row):
        b=[]
        try:
             coordinate=json.loads(requests.get(GAODE_API.format(address)).text)
             geocodes=coordinate["geocodes"][0]
             citycode=geocodes['citycode']
             adcode=geocodes['adcode']
             b.append(citycode)
             b.append(adcode)
             co=self.column+1
             for i in range(len(b)):
                self.write_excel(self.path,row,co,b[i])
                co+=1

        except Exception as e:
            print(e)
    def runall(self,row):
        key=self.read_excel()
        for i in key:
            address=i[0].value
            self.city_ad(address,row)
            row+=1




class gaode(object):
    def __init__(self,path):
        self.path=path
        self.wb = load_workbook(path)
        self.wls =self.wb['Sheet0']
        self.row=self.wls.max_row #行数
        self.column=self.wls.max_column #列数
    def read_excel(self, dir):#读取excel的地址字段
        column = input("请输入在那一列例如A：")
        address =self.wls[column+'1':column+'10'] #有多少行写多少个
        return address
    def write_excel(self, dir, row, co, body):
        try: #row代表行,column代表列
            wb=load_workbook(dir)
            wls=wb['Sheet0']
            wls.cell(row = row, column= co, value=body)
            t=wls.cell(row = row, column= co)
            wb.save(dir)
            print (t.value)
        except Exception as e:
            print(e)
    def gaode_zhuanhuan(self,dir,address,row):
        b=[]
        try:
            coordinate=json.loads(requests.get(GAODE_API.format(address)).text)
            location=coordinate["geocodes"][0]['location'].split(',')
            lng=location[0]
            lat=location[1]
            b.append(lng)
            b.append(lat)
            co=self.column+1
            for i in range(len(b)):
                self.write_excel(dir,row,co,b[i])
                co+=1

        except Exception as e:
            print(e)
    def runall(self,row):
        key=self.read_excel(self.path)
        for i in key:
            address=i[0].value
            self.gaode_zhuanhuan(self.path,address,row)
            row+=1
        
class wsg(object):
    def __init__(self,path):
        self.path=path
        self.wb = load_workbook(self.path)
        self.wls =self.wb['Sheet0']
        self.row=self.wls.max_row #行数
        self.column=self.wls.max_column #列数
        self.pi = 3.1415926535897932384626  # π
        self.a = 6378245.0  # 长半轴
        self.ee = 0.00669342162296594323  # 扁率
    def read_excel(self,dir): #读取excel内容
        column_lng=chr(ord('A')+self.column-2)
        column_lat=chr(ord('A')+self.column-1)
        glng=self.wls[column_lng+"1":column_lng+"10"]
        glat=self.wls[column_lat+"1":column_lat+"10"]
        return glng,glat
    def write_excel(self, dir, row, co, body): #row代表行,column代表列
        wb=load_workbook(dir)
        wls=wb['Sheet0']
        wls.cell(row = row, column= co, value=body)
        t=wls.cell(row = row, column= co)
        wb.save(dir)
        print (t.value)
    def transformlat(self,lng, lat):
        ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + \
            0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
        ret += (20.0 * math.sin(6.0 * lng * self.pi) + 20.0 *math.sin(2.0 * lng * self.pi)) * 2.0 / 3.0
        ret += (20.0 * math.sin(lat * self.pi) + 40.0 *math.sin(lat / 3.0 * self.pi)) * 2.0 / 3.0
        ret += (160.0 * math.sin(lat / 12.0 * self.pi) + 320 *math.sin(lat * self.pi / 30.0)) * 2.0 / 3.0
        return ret
    def transformlng(self,lng, lat):
        ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + \
            0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
        ret += (20.0 * math.sin(6.0 * lng * self.pi) + 20.0 *math.sin(2.0 * lng * self.pi)) * 2.0 / 3.0
        ret += (20.0 * math.sin(lng * self.pi) + 40.0 *math.sin(lng / 3.0 * self.pi)) * 2.0 / 3.0
        ret += (150.0 * math.sin(lng / 12.0 * self.pi) + 300.0 *math.sin(lng / 30.0 * self.pi)) * 2.0 / 3.0
        return ret
    def gcj02towgs84(self,lng, lat):
        """
        GCJ02(火星坐标系)转GPS84
        :param lng:火星坐标系的经度
        :param lat:火星坐标系纬度
        :return:
        """
        # if out_of_china(lng, lat):
        #     return lng, lat
        dlat = self.transformlat(lng - 105.0, lat - 35.0)
        dlng = self.transformlng(lng - 105.0, lat - 35.0)
        radlat = lat / 180.0 * self.pi
        magic = math.sin(radlat)
        magic = 1 - self.ee * magic * magic
        sqrtmagic = math.sqrt(magic)
        dlat = (dlat * 180.0) / ((self.a * (1 - self.ee)) / (magic * sqrtmagic) * self.pi)
        dlng = (dlng * 180.0) / (self.a / sqrtmagic * math.cos(radlat) * self.pi)
        mglat = lat + dlat
        mglng = lng + dlng
        return [lng * 2 - mglng, lat * 2 - mglat]
    def runall(self):
        zuobiao=self.read_excel(self.path)
        zipped=zip(zuobiao[0],zuobiao[1])
        row=1
        for i,j in zipped:
            try:
                b=[]
                slng=float(i[0].value)
                slat=float(j[0].value)
                lng,lat=self.gcj02towgs84(slng,slat)
                b.append(lng)
                b.append(lat)
                co=self.column+1
                for k in range(len(b)):
                    self.write_excel(self.path,row,co,b[k])
                    co+=1
                row+=1
            except Exception as e:
                print(e)
                row+=1
    



        



if __name__ == '__main__':
    # key=read_excel(path)
    # row=1
    # co=2
    # for i in key:
    #     address=i[0].value
    #     gaode_zhuanhuan(path,address,row,co)
    #     row+=1
    district_s=district(path)
    district_s.runall(1)
    gaodecode=gaode_code(path)
    gaodecode.runall(1)
    gaode_zhuan = gaode(path)
    gaode_zhuan.runall(1)#地址转换高德坐标
    wsg84=wsg(path)
    wsg84.runall()#高德坐标转换成地球坐标
